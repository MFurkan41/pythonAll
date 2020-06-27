from ipaddress import IPv4Address  # for your IP address
from pyairmore.request import AirmoreSession  # to create an AirmoreSession
from pyairmore.services.messaging import MessagingService  # to send messages
import pandas as pd
from pandas import ExcelWriter
from pandas import ExcelFile
from time import sleep as bekle
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

def printm(str1):
    for i in str1:
        print(i ,end="")
        bekle(0.03)
    print(end="\n")

ip = IPv4Address("192.168.1.4")  
session = AirmoreSession(ip)
was_accepted = session.request_authorization()
service = MessagingService(session)

messages = service.fetch_message_history()
cikart_phone = []
cikart_content = []

for i in range(0,len(messages)):
    if(messages[i].was_read == False):
        cikart_phone.append(str(messages[i].phone))
        cikart_content.append(str(messages[i].content))

filepath = "C:/Users/mfurk/Desktop/SMS GÖNDERME/deneme.xlsx"

xl = pd.ExcelFile(filepath)
df = xl.parse("Merhaba", header=None)
df.columns = ['A','B']
for i in cikart_phone:
    if(i[0] != "+"):
        a = cikart_phone.index(str(i))
        cikart_phone.pop(a)
        cikart_content.pop(a)

wb = load_workbook(filepath)
ws = wb['Merhaba']
x = ws.max_row
y = ws.max_column
deneme=list()
deneme = cikart_content

for g in range(0,len(deneme)):
    print(deneme[g])
    a = str(input("Ok:"))
    if(a==""):
        excel_sil = cikart_phone[g]
        excel_sil = excel_sil[1:] #Burayı sonra değiştirmeyi unutma. Numaralar "+90555...." den "555...." a değişecek. [3:] yapacaksın.
        for m in range(1,ws.max_row+1):
            obj = ws.cell(row=m, column=2)
            if(str(obj.value) == str(excel_sil)): #Burayı da aynı şekilde değiştirmeyi unutma.
                ws.delete_rows(m,1)
                wb.save("C:/Users/mfurk/Desktop/SMS GÖNDERME/deneme.xlsx")
    else:
        pass

