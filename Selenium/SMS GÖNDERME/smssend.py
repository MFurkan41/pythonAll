from ipaddress import IPv4Address  # for your IP address
from pyairmore.request import AirmoreSession  # to create an AirmoreSession
from pyairmore.services.messaging import MessagingService  # to send messages
from openpyxl import load_workbook
from time import sleep as bekle

def printm(str1):
    for i in str1:
        print(i ,end="")
        bekle(0.03)
    print(end="\n")




ip = IPv4Address("192.168.1.4")  
session = AirmoreSession(ip)
was_accepted = session.request_authorization()
service = MessagingService(session)

messages = service.fetch_message_history()


filepath = "C:/Users/mfurk/Desktop/SMS GÖNDERME/deneme.xlsx"
name_column = "A"
phone_column = "B"

workbook = load_workbook(filename=filepath, read_only=True)
worksheet = workbook.active

phone_numbers = []
names = []
for i in range(1000000):
    cell = "{}{}".format(phone_column, i+1)
    number = worksheet[cell].value
    if number == "None" or number is None:
        break
    if number != "" or number is not None:
        phone_numbers.append(str(number))
for a in range(1000000):
    cell = "{}{}".format(name_column, a+1)
    name = worksheet[cell].value
    if name == "None" or name is None:
        break
    if name != "" or name is not None:
        names.append(str(name))

if(len(names) == len(phone_numbers)):
    for j in range(len(names)):
        message = "Sayın {}, korona sürecinde 3000 TL'lik nakliyat sitenizi UCRETSIZ/BEDAVA kuruyoruz. Fırsatı kaçırmayın! Google'da yerinizi almak için acele edin! Örnek site: www.edirneevdenevenakliye.com Sadece masraflar size aittir. Sms Iptali için 'LISTECIKART' yazarak geri dönünüz.".format(names[j])
        #message = "Sayın {}, korona sürecinde 1 haftaligina emlak sitenizi UCRETSIZ/BEDAVA kuruyoruz. Örnek site:www.gunesemlak.com Sadece masraflar size aittir.".format(names[j])
        number = "90" + str(phone_numbers[j])
        number = "+" + str(number)
        service.send_message(number,message)
        messages = service.fetch_message_history()
        print(str(j+1)+ ". mesaj atıldı.")
else:           
    print("Excel dosyasına göre isim ve numaralar eşit sayıda değildir. Düzeltip tekrar deneyiniz..")
printm("-------------------")
printm("Tüm mesajlar gönderildi. Toplam: " + str(len(names)))
printm("Mesajlar kontrol ediliyor.")
bekle(20)
printm("-------------------")

filepath = "C:/Users/mfurk/Desktop/SMS GÖNDERME/Kapalı Telefonlar.xlsx"

workbook = load_workbook(filename=filepath, read_only=False)
worksheet = workbook.active

for e in range(len(names),0,-1):
    mesajs = 1 + int(e) - (len(names))
    messages = service.fetch_message_history()
    if(str(messages[mesajs].type) == "MessageType.SENT"):
        print(str(mesajs) + ". mesaj doğrulandı.")
    else:
        print(str(mesajs) + ". mesaj doğrulanamadı!!!")
        excelc = "A"
        excelc += str(mesajs)
        worksheet[excelc] = str(messages[e].phone)

workbook.save(filepath)

